import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
import os
from time import time
from datetime import datetime
from openpyxl import Workbook, load_workbook



# main finction
def main_func(result):
    max_range = [] ; max_range_perc = [] ; max_range_perc_10 =[]
    OH = [] ; OH_perc = []
    OL = [] ; OL_perc = []
    OC = [] ; OC_perc = [] ; OC_perc_10 = []
    for i in range(len(result['Date'])):
        max_range_ele = result['High'][i] - result['Low'][i]
        max_range.append(round(max_range_ele, 0))

        max_range_perc_ele = round( (max_range_ele / result['Open'][i])*100, 1)
        max_range_perc.append(max_range_perc_ele)

        max_range_perc_10.append(1 if max_range_perc_ele > 10 else 0)
        # -----------------------------------------------------------------
        OH_ele = result['High'][i] - result['Open'][i]
        OH.append(round(OH_ele, 0))

        OH_perc_ele = round( (OH_ele / result['Open'][i])*100, 1)
        OH_perc.append(OH_perc_ele)
        # -----------------------------------------------------------------
        OL_ele = result['Open'][i] - result['Low'][i]
        OL.append(round(OL_ele, 0))

        OL_perc_ele = round( (OL_ele / result['Open'][i])*100, 1)
        OL_perc.append(OL_perc_ele)
        # -----------------------------------------------------------------
        OC_ele = result['Open'][i] - result['Close'][i]
        OC.append(abs(round(OC_ele, 0)))

        OC_perc_ele = (OC_ele / result['Open'][i])*100
        OC_perc.append(abs( round(OC_perc_ele, 1) ))

        OC_perc_10.append(1 if abs(OC_perc_ele) > 10 else 0)
        # -----------------------------------------------------------------
        
    data_dict = {'max_range':max_range, 'max_range_perc':max_range_perc, 'max_range_perc_10':max_range_perc_10, 'OH':OH, 'OH_perc':OH_perc, 'OL':OL, 'OL_perc':OL_perc, 'OC':OC, 'OC_perc':OC_perc, 'OC_perc_10':OC_perc_10}
    result_op = pd.DataFrame(data_dict)

    return result_op


def dt_conv(date_string):
    try:
        date_object = datetime.strptime(date_string, '%Y-%m-%d')
        formatted_date = date_object.strftime('%B-%Y')
    except Exception as e:
        # print(e)
        formatted_date = ''
    return formatted_date


def run_func(file_name):
    # Fy 2021-22 -------------------------------------------------
    df_inp = pd.read_csv(file_name)
    match_values = df_inp['Date'][0:12]
    result = df_inp[df_inp['Date'].isin(match_values)]
    df_op = main_func(result)
    df_0_12 = ( pd.concat([result, df_op], axis=1) )#.drop('index', axis=1)
    new_row = ['', '', '', '', '', '', '', '', '', 
            sum(df_0_12['max_range_perc_10']), 
            '', '', '', '', '', '', 
            sum(df_0_12['OC_perc_10'])]
    df_0_12.loc[len(df_0_12)] = new_row
    df_0_12['Date'] = df_0_12['Date'].apply(dt_conv)
    # df_0_12
    # Fy 2022-23 -------------------------------------------------
    df = pd.read_csv(file_name)
    match_values = df['Date'][12:24]
    result = (df[df['Date'].isin(match_values)]).reset_index()
    df_op = main_func(result)
    df_12_24 = ( pd.concat([result, df_op], axis=1) ).drop('index', axis=1)
    new_row = ['', '', '', '', '', '', '', '', '', 
            sum(df_12_24['max_range_perc_10']), 
            '', '', '', '', '', '', 
            sum(df_12_24['OC_perc_10'])]
    df_12_24.loc[len(df_12_24)] = new_row
    df_12_24['Date'] = df_12_24['Date'].apply(dt_conv)
    # df_12_24
    # Fy 2023-24 -------------------------------------------------
    df_inp = pd.read_csv(file_name)
    match_values = df_inp['Date'][24:36]
    result = (df[df['Date'].isin(match_values)]).reset_index()
    df_op = main_func(result)
    df_24_36 = ( pd.concat([result, df_op], axis=1) ).drop('index', axis=1)
    new_row = ['', '', '', '', '', '', '', '', '', 
            sum(df_24_36['max_range_perc_10']), 
            '', '', '', '', '', '', 
            sum(df_24_36['OC_perc_10'])]
    df_24_36.loc[len(df_24_36)] = new_row
    df_24_36['Date'] = df_24_36['Date'].apply(dt_conv)
    # df_24_36
    # df_0_12, df_12_24, df_24_36
    empty_rows = pd.DataFrame([[''] * len(df_0_12.columns)] * 2, columns=df_0_12.columns)
    df_joined = pd.concat([df_0_12, empty_rows, df_12_24, empty_rows, df_24_36], ignore_index=True)

    return df_joined

def save_dfs_to_excel(df_inp, file_name, sheet_name_inp):
    try:
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            df_inp.to_excel(writer, sheet_name=sheet_name_inp)
    except Exception as e:
        print(f"An error occurred: {e}")

# def save_dfs_to_excel(dataframes, file_name, sheet_name_inp):
#     with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
#         for i, df in enumerate(dataframes):
#             # Write each dataframe to the same sheet starting at different rows
#             df.to_excel(writer, sheet_name=sheet_name_inp, startrow=i * (len(df) + 2))


def highlights(sheet_name_1):
    wb = openpyxl.load_workbook("multiple_dfs.xlsx")
    ws = wb[sheet_name_1]
    colors = [3, 3, 3, 3, 3, 3] 
    fillers = []
    for color in colors:
        temp = PatternFill(patternType='solid',
                        fgColor=Color(indexed=color))
        fillers.append(temp)
    cell_ids = ['K14', 'K29', 'K44', 'R14', 'R29', 'R44']

    for i in range(6):
        ws[cell_ids[i]].fill = fillers[i]

    wb.save("multiple_dfs.xlsx")

def ten_per_cross_count():
    sheet_name_lst = (pd.ExcelFile('multiple_dfs.xlsx')).sheet_names
    max_rng_lst_3fe = []
    oc_perc_lst_3fe = []
    tickers_3fe = []
    for i in sheet_name_lst[1:]:
        try:
            data_test = pd.read_excel('multiple_dfs.xlsx', index_col=0, sheet_name=i)
            # print('data: ', data_test)
            max_range_perc_10_lst = [data_test['max_range_perc_10'][i] for i in [12, 27, 42]]
            OC_perc_10_lst = [data_test['OC_perc_10'][i] for i in [12, 27, 42]]
            print('\n                                  Ticker: ', i)                    ; tickers_3fe.append(i)
            print('Max range percentage crossed 10% per FY :', max_range_perc_10_lst)   ; max_rng_lst_3fe.append(max_range_perc_10_lst)
            print('Open Close percentage crossed 10% per FY: ', OC_perc_10_lst)         ; oc_perc_lst_3fe.append(OC_perc_10_lst)
        except Exception as e:
            print('-------------- ERROR RLK: ', e)
    
    # print('----tickers_3fe:', tickers_3fe)
    # print('----max_rng_lst_3fe: ', max_rng_lst_3fe)
    # print('----oc_perc_lst_3fe: ', oc_perc_lst_3fe)
    new_combo_lst = []
    for i,j,k in zip(tickers_3fe, max_rng_lst_3fe, oc_perc_lst_3fe):
        new_combo_lst.append([i]+j+k)

    df_output = pd.DataFrame(new_combo_lst,
                    columns=  ['Tickers', 'Max range % > 10 FY2021-22', 'oc % > 10 FY2021-22', 'Max range % > 10 FY2022-23', 
                                'oc % > 10 FY2022-23', 'Max range % > 10 FY2023-24', 'oc % > 10 FY2023-24'])

    return df_output

def create_file(file_name):
    if os.path.exists(file_name):
        os.remove(file_name)
    wb = Workbook()
    wb.save(file_name)
    print('Excel File Created')


if __name__ == '__main__':
    t1 = time()
    file_name = 'multiple_dfs.xlsx'
    create_file(file_name)
    files = os.listdir('data/')
    for file_name in files:
        print('Files processed: ', file_name)
        # file_name = 'INFY.NS.csv'
        sheet_name_1 = ''.join( file_name.split('.')[:-1] )
        # print('sheet_name_1: ', sheet_name_1)
        file_path = 'data/' + file_name
        df_joined = run_func(file_path)
        # print('df_joined: ', df_joined)
        save_dfs_to_excel(df_joined, 'multiple_dfs.xlsx', sheet_name_1)
        highlights(sheet_name_1)

    df_output = ten_per_cross_count()
    save_dfs_to_excel(df_output, 'multiple_dfs.xlsx', 'Summary')
    
    print('\nTotal time taken (sec): ', round(time()-t1, 3) )
    print(input('Click Enter to exit'))
